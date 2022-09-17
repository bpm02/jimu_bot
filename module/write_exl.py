"""
エクセルに記入
"""
from openpyxl import load_workbook
import PyPDF2

def write_exl(INVOICE_ON, FILE, SHEET, CLIENT, ITEMS, DAYZ, FINISHED_DIR, file_name):
    if INVOICE_ON == 1:
        wb = load_workbook(FILE)
        ws = wb[SHEET]

        for k, v in CLIENT.items():
          ws[k] = v

        for i in ITEMS:
            for k, v in i.items():
                ws[k] = v

        ws['AH3'] = DAYZ # Time

        # create PdfFileWriter object
        writer = PyPDF2.PdfFileWriter()

        wb.save(FINISHED_DIR + '請求書' + file_name + '.xlsx')
